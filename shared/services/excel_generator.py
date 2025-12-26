from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

class ExcelGenerator:
    def generate(self, topic_id: int, questions: list) -> bytes:
        """Generate Excel file from questions data"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Topic {topic_id}"
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = ["ID", "Question", "Difficulty", "Expected Answer", "Tags"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row_idx, question in enumerate(questions, 2):
            ws.cell(row=row_idx, column=1, value=question.get('id'))
            ws.cell(row=row_idx, column=2, value=question.get('text'))
            ws.cell(row=row_idx, column=3, value=question.get('difficulty'))
            ws.cell(row=row_idx, column=4, value=question.get('expected_answer'))
            ws.cell(row=row_idx, column=5, value=question.get('tags'))
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 30
        
        # Save to bytes
        excel_stream = BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        
        return excel_stream.getvalue()